import os

from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned
from django.core.management import BaseCommand
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

from authority.models import Country
from controlled_list.models import Nationality
from research.models import Researcher, ResearcherDegree


class Command(BaseCommand):
    def handle(self, *args, **options):
        filename = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'ResearchRequests.xlsx')
        wb = load_workbook(filename=filename)
        self.create_researchers(wb)
        # self.add_researcher_extra_data(wb)

    def get_value(self, row, index):
        if isinstance(row[index].value, str):
            return row[index].value.strip() if row[index].value else ''
        else:
            return row[index].value

    def create_researchers(self, wb):
        researchers = wb['Researcher - Base Data']
        for row in researchers.iter_rows(min_row=2, max_row=215):
            last_name = self.get_value(row, 0)
            first_name = self.get_value(row, 1)
            middle_name = self.get_value(row, 2)
            card_no = self.get_value(row, 4)
            passport = row[5].value if row[5].value else "AA111111"
            email = self.get_value(row, 6)
            address_hungary = self.get_value(row, 7)
            city_hungary = self.get_value(row, 8)
            address_abroad = self.get_value(row, 9)
            city_abroad = self.get_value(row, 11)
            country_abroad = self.get_value(row, 12)
            registration_date = self.get_value(row, 13)

            if last_name:
                try:
                    researcher = Researcher.objects.get(
                        card_number=card_no
                    )
                    print("Researcher: %s is updated!" % researcher.name)
                except ObjectDoesNotExist:
                    try:
                        researcher = Researcher.objects.get(
                            first_name=first_name,
                            last_name=last_name
                        )
                        print("Researcher: %s is updated!" % researcher.name)
                    except ObjectDoesNotExist:
                        researcher = Researcher.objects.create(
                            first_name=first_name,
                            last_name=last_name,
                            middle_name=middle_name,
                            card_number=card_no,
                            id_number=passport,
                            email=email,
                            address_hungary=address_hungary,
                            address_abroad=address_abroad,
                            city_hungary=city_hungary,
                            city_abroad=city_abroad,
                            date_created=registration_date
                        )
                        print("Researcher: %s is created!" % researcher.name)
                    except MultipleObjectsReturned:
                        print("Multiple researchers with name are registered: %s %s" % (first_name, last_name))
                        researcher = Researcher.objects.filter(
                            first_name=first_name,
                            last_name=last_name
                        ).first()
                except MultipleObjectsReturned:
                    print("Multiple card numbers are registered: %s" % card_no)
                    researcher = Researcher.objects.filter(
                        card_number=card_no
                    ).first()

                try:
                    country = Country.objects.get(
                        country=country_abroad
                    )
                    researcher.country = country
                    researcher.save()
                except ObjectDoesNotExist:
                    pass

    def add_researcher_extra_data(self, wb):
        topics = wb['Researcher - Topics']
        for row in topics.iter_rows(min_row=2):
            researcher = self.get_value(row, 0)
            research_subject = self.get_value(row, 1)
            status = self.get_value(row, 2)
            status_other = self.get_value(row, 3)
            degree = self.get_value(row, 4)
            employer = self.get_value(row, 6)
            occupation = self.get_value(row, 7)
            nationality = self.get_value(row, 8)

            try:
                researcher = Researcher.objects.get(
                    name=researcher
                )
            except ObjectDoesNotExist:
                print("Can't find researcher in topics tab %s" % researcher)
                return
            except MultipleObjectsReturned:
                print("Multiple objects returned for %s" % researcher)
                return

            researcher.research_subject = research_subject
            researcher.employer_or_school = employer

            try:
                d = ResearcherDegree.objects.get(degree=degree)
                researcher.degree = d
            except ObjectDoesNotExist:
                pass

            try:
                n = Nationality.objects.get(nationality=nationality)
                researcher.nationality = n
            except ObjectDoesNotExist:
                pass